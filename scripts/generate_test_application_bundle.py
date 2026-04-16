#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from copy import deepcopy
from ipaddress import ip_address
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ENV_FILE = (
    REPO_ROOT
    / "deploy-all"
    / "qwenpaw"
    / "working"
    / "workspaces"
    / "query"
    / "skills"
    / "veops-cmdb"
    / ".env"
)
DEFAULT_OUTPUT_DIR = REPO_ROOT / "portal" / "test-fixtures" / "test-zhiguan-clone"
MODEL_SHEET_ORDER = [
    "project",
    "vserver",
    "docker",
    "database",
    "mysql",
    "PostgreSQL",
    "redis",
    "Kafka",
    "elasticsearch",
]
PREFERRED_FIELD_ORDER = [
    "project_name",
    "dev_no",
    "dev_name",
    "vserver_name",
    "private_ip",
    "manage_ip",
    "middleware_name",
    "middleware_ip",
    "db_instance",
    "db_ip",
    "db_port",
    "middleware_port",
    "property_no",
    "platform",
    "AssociatedVM",
    "AssociatedPhyMachine",
    "host_name",
    "project_type",
    "project_status",
    "project_description",
    "dev_model",
    "dev_class",
    "vendor",
    "server_room",
    "data_center",
    "city",
    "county",
    "op_duty",
]
EXCLUDED_EXPORT_FIELDS = {
    "_id",
    "_type",
    "_updated_at",
    "_updated_by",
    "ci_type_alias",
    "unique",
    "unique_alias",
}
IP_FIELDS = {
    "manage_ip",
    "middleware_ip",
    "db_ip",
    "private_ip",
    "manager_ip",
    "host_ip",
    "AssociatedVM",
    "AssociatedPhyMachine",
}
ID_SUFFIX_FIELDS = {
    "property_no",
    "db_instance",
    "middleware_name",
}
PREFIX_FIELDS = {
    "dev_no",
    "dev_name",
}


def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("'").strip('"')
    return values


def login(base_url: str, username: str, password: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({"Accept-Language": "zh"})
    response = session.post(
        f"{base_url.rstrip('/')}/api/v1/acl/login",
        json={"username": username, "password": password},
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    token = payload.get("token")
    if not token:
        raise RuntimeError("登录成功但响应中缺少 token")
    session.headers["Access-Token"] = token
    return session


def fetch_all_by_platform(
    session: requests.Session,
    *,
    base_url: str,
    platform_name: str,
) -> list[dict[str, Any]]:
    page = 1
    rows: list[dict[str, Any]] = []
    while True:
        response = session.get(
            f"{base_url.rstrip('/')}/api/v0.1/ci/s",
            params={
                "q": f"platform:*{platform_name}*",
                "count": 100,
                "page": page,
            },
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        batch = payload.get("result") or []
        rows.extend(batch)
        total = int(payload.get("numfound") or len(batch))
        if not batch or len(rows) >= total:
            return rows
        page += 1


def fetch_projects(session: requests.Session, *, base_url: str) -> list[dict[str, Any]]:
    response = session.get(
        f"{base_url.rstrip('/')}/api/v0.1/ci/s",
        params={"q": "_type:3", "count": 100, "page": 1},
        timeout=30,
    )
    response.raise_for_status()
    return response.json().get("result") or []


def fetch_relation_children(
    session: requests.Session,
    *,
    base_url: str,
    root_id: int,
) -> list[dict[str, Any]]:
    response = session.get(
        f"{base_url.rstrip('/')}/api/v0.1/ci_relations/s",
        params={"root_id": root_id, "level": 1, "reverse": 0, "count": 10000},
        timeout=30,
    )
    response.raise_for_status()
    return response.json().get("result") or []


def fetch_application_topology_rows(
    session: requests.Session,
    *,
    base_url: str,
    app_name: str,
) -> list[dict[str, Any]]:
    projects = fetch_projects(session, base_url=base_url)
    project = next((row for row in projects if str(row.get("project_name") or "").strip() == app_name), None)
    if not project:
        raise RuntimeError(f"未找到应用 project_name={app_name}")

    rows_by_id: dict[int, dict[str, Any]] = {}

    def remember(row: dict[str, Any]) -> None:
        row_id = row.get("_id")
        if isinstance(row_id, int) and row_id not in rows_by_id:
            rows_by_id[row_id] = row

    remember(project)
    level1 = fetch_relation_children(session, base_url=base_url, root_id=project["_id"])
    for row in level1:
        remember(row)

    vservers = [row for row in rows_by_id.values() if str(row.get("ci_type") or "") == "vserver"]
    for vserver in vservers:
        for row in fetch_relation_children(session, base_url=base_url, root_id=vserver["_id"]):
            remember(row)

    dockers = [row for row in rows_by_id.values() if str(row.get("ci_type") or "") == "docker"]
    for docker in dockers:
        for row in fetch_relation_children(session, base_url=base_url, root_id=docker["_id"]):
            remember(row)

    return list(rows_by_id.values())


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(item) for item in value if item not in (None, ""))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def looks_like_ip(value: str) -> bool:
    try:
        ip_address(value)
        return True
    except ValueError:
        return False


def remap_ip(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return raw
    if not looks_like_ip(raw):
        return raw
    octets = raw.split(".")
    if octets[:2] == ["10", "43"]:
        return f"10.253.{octets[2]}.{octets[3]}"
    if octets[:3] == ["172", "27", "34"]:
        return f"172.31.134.{octets[3]}"
    if octets[:3] == ["172", "28", "75"]:
        return f"172.31.75.{octets[3]}"
    if octets[:2] == ["198", "18"]:
        return f"198.19.{octets[2]}.{octets[3]}"
    return raw


def clone_text_identifier(value: str, *, prefix: str = "TESTZG-", suffix: str = "-testzg") -> str:
    raw = str(value or "").strip()
    if not raw:
        return raw
    if looks_like_ip(raw):
        return remap_ip(raw)
    if raw.startswith("host_network_"):
        return f"{raw}_testzg"
    if re.fullmatch(r"\d+", raw):
        return raw
    if raw.endswith(suffix):
        return raw
    if raw.startswith(prefix):
        return raw
    return f"{prefix}{raw}"


def transform_scalar(field: str, value: Any, *, source_app: str, target_app: str) -> Any:
    if value is None:
        return None
    if field == "platform":
        return [target_app]
    if field == "project_name":
        return target_app
    if field == "p_id":
        try:
            return int(value) + 500000
        except Exception:
            return f"{value}-testzg"
    if field in IP_FIELDS:
        return remap_ip(str(value))
    if field in PREFIX_FIELDS:
        return clone_text_identifier(str(value), prefix="TESTZG-")
    if field in ID_SUFFIX_FIELDS:
        raw = str(value).strip()
        if not raw:
            return raw
        if field == "middleware_name" and looks_like_ip(raw):
            return remap_ip(raw)
        return f"{raw}-testzg"
    if field == "vserver_name":
        raw = str(value).strip()
        if not raw:
            return raw
        return remap_ip(raw) if looks_like_ip(raw) else f"TESTZG-{raw}"
    if isinstance(value, str):
        return value.replace(source_app, target_app)
    return value


def transform_row(row: dict[str, Any], *, source_app: str, target_app: str) -> dict[str, Any]:
    next_row = deepcopy(row)
    for field, value in list(next_row.items()):
        if field == "platform":
            next_row[field] = [target_app]
        elif isinstance(value, list):
            next_row[field] = [
                transform_scalar(field, item, source_app=source_app, target_app=target_app)
                for item in value
            ]
        else:
            next_row[field] = transform_scalar(field, value, source_app=source_app, target_app=target_app)
    return next_row


def unique_value(row: dict[str, Any]) -> str:
    key = str(row.get("unique") or "").strip()
    value = row.get(key)
    if isinstance(value, list):
        return stringify(value[:1]).strip()
    return stringify(value).strip()


def sheet_title(ci_type: str, alias: str) -> str:
    label = alias or ci_type or "Sheet"
    return label[:31]


def ordered_fields(rows: list[dict[str, Any]]) -> list[str]:
    field_set: set[str] = set()
    for row in rows:
        field_set.update(key for key in row if key not in EXCLUDED_EXPORT_FIELDS and not key.startswith("_"))
    preferred = [field for field in PREFERRED_FIELD_ORDER if field in field_set]
    remaining = sorted(field_set - set(preferred))
    return preferred + remaining


def build_relation_rows(rows: list[dict[str, Any]], *, app_name: str) -> list[dict[str, Any]]:
    project_row = next((row for row in rows if row.get("ci_type") == "project"), None)
    relations: list[dict[str, Any]] = []
    def append_relation(
        *,
        source_model: str,
        source_unique_field: str,
        source_unique_value: str,
        relation_type: str,
        target_model: str,
        target_unique_field: str,
        target_unique_value: str,
        evidence: str,
    ) -> None:
        if not all(
            [
                source_model,
                source_unique_field,
                source_unique_value,
                relation_type,
                target_model,
                target_unique_field,
                target_unique_value,
            ]
        ):
            return
        relations.append(
            {
                "source_model": source_model,
                "source_unique_field": source_unique_field,
                "source_unique_value": source_unique_value,
                "relation_type": relation_type,
                "target_model": target_model,
                "target_unique_field": target_unique_field,
                "target_unique_value": target_unique_value,
                "evidence": evidence,
            }
        )

    vserver_by_ip: dict[str, dict[str, Any]] = {}
    docker_by_ip: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("ci_type") != "vserver":
            continue
        private_ip = row.get("private_ip")
        if isinstance(private_ip, list):
            for item in private_ip:
                if item:
                    vserver_by_ip[str(item)] = row
        elif private_ip:
            vserver_by_ip[str(private_ip)] = row

    for row in rows:
        if row.get("ci_type") != "docker":
            continue
        manage_ip = stringify(row.get("manage_ip")).strip()
        middleware_name = stringify(row.get("middleware_name")).strip()
        if manage_ip:
            docker_by_ip[manage_ip] = row
        if middleware_name and looks_like_ip(middleware_name):
            docker_by_ip[middleware_name] = row

    for row in rows:
        ci_type = row.get("ci_type")
        if not ci_type or ci_type == "project":
            continue
        if project_row and ci_type not in {"vserver", "docker", "networkdevice"}:
            append_relation(
                source_model="project",
                source_unique_field="project_name",
                source_unique_value=app_name,
                relation_type="contain",
                target_model=ci_type,
                target_unique_field=row.get("unique"),
                target_unique_value=unique_value(row),
                evidence="platform 字段属于同一应用",
            )

    for row in rows:
        if row.get("ci_type") != "vserver":
            continue
        append_relation(
            source_model="project",
            source_unique_field="project_name",
            source_unique_value=app_name,
            relation_type="deploy",
            target_model="vserver",
            target_unique_field=row.get("unique"),
            target_unique_value=unique_value(row),
            evidence="应用运行在该虚拟机上",
        )

    for row in rows:
        if row.get("ci_type") != "docker":
            continue
        associated_vm = stringify(row.get("AssociatedVM")).strip()
        vserver = vserver_by_ip.get(associated_vm)
        if not vserver:
            continue
        append_relation(
            source_model="vserver",
            source_unique_field=vserver.get("unique"),
            source_unique_value=unique_value(vserver),
            relation_type="connect",
            target_model="docker",
            target_unique_field=row.get("unique"),
            target_unique_value=unique_value(row),
            evidence="docker.AssociatedVM 指向虚拟机内网 IP",
        )

    for row in rows:
        ci_type = row.get("ci_type")
        if ci_type not in {"database", "mysql", "PostgreSQL", "redis", "Kafka", "elasticsearch"}:
            continue
        target_ip = (
            stringify(row.get("manage_ip")).strip()
            or stringify(row.get("middleware_ip")).strip()
            or stringify(row.get("db_ip")).strip()
        )
        docker = docker_by_ip.get(target_ip)
        if not docker:
            continue
        append_relation(
            source_model="docker",
            source_unique_field=docker.get("unique"),
            source_unique_value=unique_value(docker),
            relation_type="deploy",
            target_model=ci_type,
            target_unique_field=row.get("unique"),
            target_unique_value=unique_value(row),
            evidence="中间件/数据库 IP 与 docker 管理 IP 一致",
        )

    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    for row in relations:
        key = (
            row["source_model"],
            row["source_unique_field"],
            row["source_unique_value"],
            row["relation_type"],
            row["target_model"],
            row["target_unique_field"],
            row["target_unique_value"],
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv_bundle(output_dir: Path, grouped_rows: dict[tuple[str, str], list[dict[str, Any]]], relations: list[dict[str, Any]]) -> list[Path]:
    csv_dir = output_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    files: list[Path] = []
    ordered_groups = sorted(
        grouped_rows.items(),
        key=lambda item: (
            MODEL_SHEET_ORDER.index(item[0][0]) if item[0][0] in MODEL_SHEET_ORDER else 999,
            item[0][0],
        ),
    )
    for index, ((ci_type, alias), rows) in enumerate(ordered_groups, start=1):
        filename = f"{index:02d}-{ci_type}.csv"
        path = csv_dir / filename
        fields = ordered_fields(rows)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            for row in rows:
                writer.writerow({field: stringify(row.get(field)) for field in fields})
        files.append(path)

    relation_path = csv_dir / "99-relations-reference.csv"
    with relation_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = [
            "source_model",
            "source_unique_field",
            "source_unique_value",
            "relation_type",
            "target_model",
            "target_unique_field",
            "target_unique_value",
            "evidence",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in relations:
            writer.writerow(row)
    files.append(relation_path)
    return files


def write_xlsx_bundle(
    path: Path,
    grouped_rows: dict[tuple[str, str], list[dict[str, Any]]],
    relations: list[dict[str, Any]],
    *,
    source_app: str,
    target_app: str,
) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "说明"
    summary_lines = [
        ("来源应用", source_app),
        ("目标测试应用", target_app),
        ("用途", "基于真实 CMDB 数据克隆的整套导入测试样例"),
        ("改动原则", "应用名、平台归属、唯一值、关键 IP 已替换；仅保留源应用真实拓扑涉及的数据"),
        ("关系参考", "关系参考 sheet 给出应用、虚拟机、docker 到下游组件的显式关系"),
    ]
    for row_index, (key, value) in enumerate(summary_lines, start=1):
        summary_ws.cell(row=row_index, column=1, value=key)
        summary_ws.cell(row=row_index, column=2, value=value)
    summary_ws.freeze_panes = "A2"

    ordered_groups = sorted(
        grouped_rows.items(),
        key=lambda item: (
            MODEL_SHEET_ORDER.index(item[0][0]) if item[0][0] in MODEL_SHEET_ORDER else 999,
            item[0][0],
        ),
    )
    for ci_type, alias in [item[0] for item in ordered_groups]:
        rows = grouped_rows[(ci_type, alias)]
        ws = workbook.create_sheet(title=sheet_title(ci_type, alias))
        fields = ordered_fields(rows)
        for col_index, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col_index, value=field)
        for row_index, row in enumerate(rows, start=2):
            for col_index, field in enumerate(fields, start=1):
                ws.cell(row=row_index, column=col_index, value=stringify(row.get(field)))
        ws.freeze_panes = "A2"

    relation_ws = workbook.create_sheet(title="关系参考")
    relation_fields = [
        "source_model",
        "source_unique_field",
        "source_unique_value",
        "relation_type",
        "target_model",
        "target_unique_field",
        "target_unique_value",
        "evidence",
    ]
    for col_index, field in enumerate(relation_fields, start=1):
        relation_ws.cell(row=1, column=col_index, value=field)
    for row_index, row in enumerate(relations, start=2):
        for col_index, field in enumerate(relation_fields, start=1):
            relation_ws.cell(row=row_index, column=col_index, value=row.get(field))
    relation_ws.freeze_panes = "A2"

    workbook.save(path)


def write_summary_markdown(
    path: Path,
    *,
    source_app: str,
    target_app: str,
    grouped_rows: dict[tuple[str, str], list[dict[str, Any]]],
    relation_count: int,
    csv_files: list[Path],
    workbook_path: Path,
) -> None:
    lines = [
        f"# {target_app} 导入测试包",
        "",
        f"- 来源应用：`{source_app}`",
        f"- 克隆目标：`{target_app}`",
        f"- 生成时间：基于当前 CMDB 实时拉取",
        f"- Excel 包：`{workbook_path.name}`",
        "",
        "## 资源统计",
        "",
        "| 模型 | 别名 | 条数 |",
        "| --- | --- | ---: |",
    ]
    ordered_groups = sorted(
        grouped_rows.items(),
        key=lambda item: (
            MODEL_SHEET_ORDER.index(item[0][0]) if item[0][0] in MODEL_SHEET_ORDER else 999,
            item[0][0],
        ),
    )
    for (ci_type, alias), rows in ordered_groups:
        lines.append(f"| `{ci_type}` | {alias or ci_type} | {len(rows)} |")
    lines.extend(
        [
            "",
            "## 关系参考",
            "",
            f"- 自动生成关系参考 {relation_count} 条。",
            "- `contain`：应用到应用内资源，依据 `platform` 字段。",
            "- `deploy`：应用到虚拟机、虚拟机到 docker、docker 到中间件/数据库。",
            "",
            "## CSV 文件",
            "",
        ]
    )
    for file in csv_files:
        lines.append(f"- `{file.name}`")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="从现网应用生成一套可导入的测试应用包")
    parser.add_argument("--source-app", default="天翼智观")
    parser.add_argument("--target-app", default="测试智观")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV_FILE))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    args = parser.parse_args()

    env = parse_env_file(Path(args.env_file))
    base_url = env["VEOPS_BASE_URL"]
    username = env["VEOPS_USERNAME"]
    password = env["VEOPS_PASSWORD"]

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    session = login(base_url, username, password)
    source_rows = fetch_application_topology_rows(session, base_url=base_url, app_name=args.source_app)
    if not source_rows:
        raise RuntimeError(f"未找到应用 {args.source_app} 的任何拓扑数据")

    transformed_rows = [
        transform_row(row, source_app=args.source_app, target_app=args.target_app)
        for row in source_rows
    ]
    grouped_rows: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in transformed_rows:
        grouped_rows[(str(row.get("ci_type") or ""), str(row.get("ci_type_alias") or ""))].append(row)

    relations = build_relation_rows(transformed_rows, app_name=args.target_app)

    source_json_path = output_dir / "tiyizhiguan-source-export.json"
    target_json_path = output_dir / "test-zhiguan-clone.json"
    workbook_path = output_dir / "test-zhiguan-import-bundle.xlsx"
    summary_path = output_dir / "test-zhiguan-import-summary.md"

    write_json(source_json_path, source_rows)
    write_json(target_json_path, transformed_rows)
    csv_files = write_csv_bundle(output_dir, grouped_rows, relations)
    write_xlsx_bundle(
        workbook_path,
        grouped_rows,
        relations,
        source_app=args.source_app,
        target_app=args.target_app,
    )
    write_summary_markdown(
        summary_path,
        source_app=args.source_app,
        target_app=args.target_app,
        grouped_rows=grouped_rows,
        relation_count=len(relations),
        csv_files=csv_files,
        workbook_path=workbook_path,
    )

    print(
        json.dumps(
            {
                "source_count": len(source_rows),
                "target_count": len(transformed_rows),
                "relation_count": len(relations),
                "workbook": str(workbook_path),
                "summary": str(summary_path),
                "csv_dir": str(output_dir / "csv"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
